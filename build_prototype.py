import sys, io, json, os
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import LineChart, BarChart, Reference
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import ColorScaleRule
from datetime import datetime

# 데이터 로드
with open('data/raw/comtrade_284190_korea_export.json', 'r', encoding='utf-8') as f:
    raw = json.load(f)

records = sorted(raw, key=lambda x: x.get('period', ''))
records = [{
    'period': r['period'],
    'value_usd': r.get('primaryValue', 0),
    'wgt_ton': (r.get('netWgt') or 0) / 1000,
} for r in records]

for i, r in enumerate(records):
    r['mom'] = None
    r['yoy'] = None
    if i >= 1 and records[i-1]['value_usd']:
        r['mom'] = r['value_usd'] / records[i-1]['value_usd'] - 1
    if i >= 12 and records[i-12]['value_usd']:
        r['yoy'] = r['value_usd'] / records[i-12]['value_usd'] - 1
    r['unit_price'] = r['value_usd'] / r['wgt_ton'] if r['wgt_ton'] else 0

# 스타일
NAVY = '1F3864'
LIGHT_BLUE = 'D9E1F2'
GRAY = 'F2F2F2'
WHITE = 'FFFFFF'

title_font = Font(name='맑은 고딕', size=14, bold=True, color=WHITE)
title_fill = PatternFill('solid', fgColor=NAVY)
header_font = Font(name='맑은 고딕', size=10, bold=True, color=WHITE)
header_fill = PatternFill('solid', fgColor=NAVY)
sub_fill = PatternFill('solid', fgColor=LIGHT_BLUE)
gray_fill = PatternFill('solid', fgColor=GRAY)
center = Alignment(horizontal='center', vertical='center')
right = Alignment(horizontal='right', vertical='center')
left = Alignment(horizontal='left', vertical='center')
thin = Side(style='thin', color='BFBFBF')
border = Border(left=thin, right=thin, top=thin, bottom=thin)

wb = openpyxl.Workbook()

# ============ Sheet 1: Dashboard ============
ws = wb.active
ws.title = 'Dashboard'
ws.sheet_view.showGridLines = False

ws.merge_cells('B2:K2')
ws['B2'] = '수출입통계 트래커 — 종목별 대시보드'
ws['B2'].font = title_font
ws['B2'].fill = title_fill
ws['B2'].alignment = center
ws.row_dimensions[2].height = 28

info_labels = [
    ('종목명', '에코프로비엠'),
    ('종목코드', '247540'),
    ('시장', 'KOSDAQ'),
    ('시총(억원)', '93,847'),
    ('주력제품', '하이니켈 NCM/NCA 양극활물질'),
    ('HS코드', '2841.90'),
    ('HS 품목명', '산의 금속산염류 (리튬·니켈·코발트·망간 화합물 포함)'),
    ('매출비중', '100% (단일사업)'),
    ('데이터 출처', 'UN Comtrade — 한국 → 전세계 수출 (HS6 기준, FOB USD)'),
    ('데이터 기간', f"{records[0]['period']} ~ {records[-1]['period']}  ({len(records)}개월)"),
]

start_row = 4
for i, (k, v) in enumerate(info_labels):
    r = start_row + i
    ws.cell(row=r, column=2, value=k).font = Font(name='맑은 고딕', size=10, bold=True, color=NAVY)
    ws.cell(row=r, column=2).fill = sub_fill
    ws.cell(row=r, column=2).alignment = left
    ws.cell(row=r, column=2).border = border
    ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=6)
    ws.cell(row=r, column=3, value=v).font = Font(name='맑은 고딕', size=10)
    ws.cell(row=r, column=3).alignment = left
    for c in range(3, 7):
        ws.cell(row=r, column=c).border = border

# KPI
latest = records[-1]
v_latest = latest['value_usd']
v_3m_avg = sum(r['value_usd'] for r in records[-3:]) / 3
v_12m_avg = sum(r['value_usd'] for r in records[-12:]) / 12
v_peak = max(r['value_usd'] for r in records)
peak_period = next(r['period'] for r in records if r['value_usd'] == v_peak)

kpis = [
    ('최근월', f"{latest['period']}"),
    ('최근월 수출액', f"${v_latest/1e6:,.1f}M"),
    ('YoY', f"{latest['yoy']*100:+.1f}%" if latest['yoy'] is not None else 'N/A'),
    ('MoM', f"{latest['mom']*100:+.1f}%" if latest['mom'] is not None else 'N/A'),
    ('3개월 평균', f"${v_3m_avg/1e6:,.1f}M"),
    ('12개월 평균', f"${v_12m_avg/1e6:,.1f}M"),
    ('역대 최고', f"${v_peak/1e6:,.1f}M  ({peak_period})"),
    ('현재 vs 역대최고', f"{(v_latest/v_peak - 1)*100:+.1f}%"),
]
for i, (k, v) in enumerate(kpis):
    r = start_row + i
    ws.cell(row=r, column=8, value=k).font = Font(name='맑은 고딕', size=10, bold=True, color=NAVY)
    ws.cell(row=r, column=8).fill = sub_fill
    ws.cell(row=r, column=8).alignment = left
    ws.cell(row=r, column=8).border = border
    ws.merge_cells(start_row=r, start_column=9, end_row=r, end_column=11)
    ws.cell(row=r, column=9, value=v).font = Font(name='맑은 고딕', size=11, bold=True)
    ws.cell(row=r, column=9).alignment = right
    for c in range(9, 12):
        ws.cell(row=r, column=c).border = border

table_start = start_row + len(info_labels) + 2
ws.cell(row=table_start, column=2, value='월별 수출 시계열').font = Font(name='맑은 고딕', size=12, bold=True, color=NAVY)

th_row = table_start + 1
headers = ['연월', '수출액 (USD)', '수출중량 (톤)', '단가 ($/kg)', 'MoM', 'YoY']
for i, h in enumerate(headers):
    c = ws.cell(row=th_row, column=2+i, value=h)
    c.font = header_font
    c.fill = header_fill
    c.alignment = center
    c.border = border

for i, r in enumerate(records):
    rr = th_row + 1 + i
    ws.cell(row=rr, column=2, value=r['period']).alignment = center
    ws.cell(row=rr, column=3, value=r['value_usd']).number_format = '#,##0'
    ws.cell(row=rr, column=3).alignment = right
    ws.cell(row=rr, column=4, value=round(r['wgt_ton'], 1)).number_format = '#,##0.0'
    ws.cell(row=rr, column=4).alignment = right
    ws.cell(row=rr, column=5, value=r['unit_price']/1000 if r['unit_price'] else None).number_format = '#,##0.00'
    ws.cell(row=rr, column=5).alignment = right
    ws.cell(row=rr, column=6, value=r['mom']).number_format = '+0.0%;-0.0%;-'
    ws.cell(row=rr, column=6).alignment = right
    ws.cell(row=rr, column=7, value=r['yoy']).number_format = '+0.0%;-0.0%;-'
    ws.cell(row=rr, column=7).alignment = right
    if i % 2 == 0:
        for c in range(2, 8):
            ws.cell(row=rr, column=c).fill = gray_fill
    for c in range(2, 8):
        ws.cell(row=rr, column=c).border = border
        ws.cell(row=rr, column=c).font = Font(name='맑은 고딕', size=10)

last_data_row = th_row + len(records)
ws.conditional_formatting.add(
    f'G{th_row+1}:G{last_data_row}',
    ColorScaleRule(start_type='num', start_value=-0.5, start_color='F8696B',
                   mid_type='num', mid_value=0, mid_color='FFEB84',
                   end_type='num', end_value=0.5, end_color='63BE7B'))

ws.column_dimensions['A'].width = 2
ws.column_dimensions['B'].width = 14
ws.column_dimensions['C'].width = 18
ws.column_dimensions['D'].width = 14
ws.column_dimensions['E'].width = 12
ws.column_dimensions['F'].width = 10
ws.column_dimensions['G'].width = 10
ws.column_dimensions['H'].width = 16
ws.column_dimensions['I'].width = 12
ws.column_dimensions['J'].width = 12
ws.column_dimensions['K'].width = 12

# 차트
chart1 = LineChart()
chart1.title = '월별 수출액 추이 (USD)'
chart1.y_axis.title = '수출액'
chart1.x_axis.title = '연월'
chart1.height = 9
chart1.width = 22
data_ref = Reference(ws, min_col=3, min_row=th_row, max_col=3, max_row=last_data_row)
cat_ref = Reference(ws, min_col=2, min_row=th_row+1, max_row=last_data_row)
chart1.add_data(data_ref, titles_from_data=True)
chart1.set_categories(cat_ref)
chart1.legend = None
ws.add_chart(chart1, 'I16')

chart2 = BarChart()
chart2.type = 'col'
chart2.title = 'YoY 증감률'
chart2.y_axis.title = 'YoY'
chart2.x_axis.title = '연월'
chart2.height = 9
chart2.width = 22
yoy_ref = Reference(ws, min_col=7, min_row=th_row, max_col=7, max_row=last_data_row)
chart2.add_data(yoy_ref, titles_from_data=True)
chart2.set_categories(cat_ref)
chart2.legend = None
ws.add_chart(chart2, 'I35')

# ============ Sheet 2: Universe ============
src_wb = openpyxl.load_workbook('개별종목_시총1500억이상_TOP500_20260503.xlsx', data_only=True)
src_ws = src_wb.active
ws_u = wb.create_sheet('Universe')
ws_u.sheet_view.showGridLines = False

ws_u.merge_cells('A1:L1')
ws_u['A1'] = 'Universe — 시총 1,500억 이상 TOP500  (기준일: 2026.05.03)'
ws_u['A1'].font = title_font
ws_u['A1'].fill = title_fill
ws_u['A1'].alignment = center
ws_u.row_dimensions[1].height = 24

new_headers = ['순위', '종목코드', '종목명', '시장', '시총(억원)', '현재가', '거래량',
               '상장주식수(천주)', 'PER', 'ROE', 'HS 매핑상태', '편입여부']
for i, h in enumerate(new_headers):
    c = ws_u.cell(row=2, column=i+1, value=h)
    c.font = header_font
    c.fill = header_fill
    c.alignment = center
    c.border = border

for src_row in src_ws.iter_rows(min_row=3, max_row=502, values_only=True):
    if src_row[0] is None: continue
    target_row = ws_u.max_row + 1
    for i, v in enumerate(src_row):
        c = ws_u.cell(row=target_row, column=i+1, value=v)
        c.font = Font(name='맑은 고딕', size=9)
        c.border = border
        c.alignment = center if i in (0,1,3) else (right if i in (4,5,6,7,8,9) else left)
    if src_row[2] == '에코프로비엠':
        for c_idx in range(1, 11):
            ws_u.cell(row=target_row, column=c_idx).fill = PatternFill('solid', fgColor='FFF2CC')
        ws_u.cell(row=target_row, column=11, value='완료 (HS 2841.90)').font = Font(name='맑은 고딕', size=9, bold=True, color='006100')
        ws_u.cell(row=target_row, column=11).fill = PatternFill('solid', fgColor='C6EFCE')
        ws_u.cell(row=target_row, column=11).border = border
        ws_u.cell(row=target_row, column=11).alignment = center
        ws_u.cell(row=target_row, column=12, value='편입').font = Font(name='맑은 고딕', size=9, bold=True, color='006100')
        ws_u.cell(row=target_row, column=12).fill = PatternFill('solid', fgColor='C6EFCE')
        ws_u.cell(row=target_row, column=12).border = border
        ws_u.cell(row=target_row, column=12).alignment = center
    else:
        ws_u.cell(row=target_row, column=11, value='미매핑').font = Font(name='맑은 고딕', size=9, color='808080')
        ws_u.cell(row=target_row, column=11).border = border
        ws_u.cell(row=target_row, column=11).alignment = center
        ws_u.cell(row=target_row, column=12, value='-').alignment = center
        ws_u.cell(row=target_row, column=12).border = border

widths = [6, 10, 18, 8, 12, 11, 13, 14, 8, 8, 22, 10]
for i, w in enumerate(widths):
    ws_u.column_dimensions[get_column_letter(i+1)].width = w
ws_u.freeze_panes = 'A3'

# ============ Sheet 3: HS_Mapping ============
ws_m = wb.create_sheet('HS_Mapping')
ws_m.sheet_view.showGridLines = False
ws_m.merge_cells('A1:I1')
ws_m['A1'] = '종목 ↔ HS코드 매핑 마스터'
ws_m['A1'].font = title_font
ws_m['A1'].fill = title_fill
ws_m['A1'].alignment = center
ws_m.row_dimensions[1].height = 24

map_headers = ['종목코드', '종목명', 'HS6', 'HS 한글품명', '주력제품 키워드', '매출비중', '매핑근거', '신뢰도', '검토상태']
for i, h in enumerate(map_headers):
    c = ws_m.cell(row=2, column=i+1, value=h)
    c.font = header_font
    c.fill = header_fill
    c.alignment = center
    c.border = border

mapping_row = ['247540', '에코프로비엠', '2841.90',
               '산의 금속산염류 (기타)',
               'NCM/NCA 양극활물질, 리튬코발트산화물, 니켈코발트망간산리튬',
               '100%',
               'DART 사업보고서 매출구성 100% 양극재. 양극재는 HS 2841.90으로 분류.',
               '높음 (1.00)',
               '확정']
for i, v in enumerate(mapping_row):
    c = ws_m.cell(row=3, column=i+1, value=v)
    c.font = Font(name='맑은 고딕', size=10)
    c.border = border
    c.alignment = center if i in (0, 2, 5, 7, 8) else left

widths_m = [10, 14, 8, 22, 32, 9, 40, 11, 10]
for i, w in enumerate(widths_m):
    ws_m.column_dimensions[get_column_letter(i+1)].width = w
ws_m.row_dimensions[3].height = 30

# ============ Sheet 4: Raw_Data ============
ws_r = wb.create_sheet('Raw_Data')
ws_r.sheet_view.showGridLines = False
ws_r.merge_cells('A1:F1')
ws_r['A1'] = '원본 데이터 — UN Comtrade (HS 2841.90, 한국 -> 전세계 수출)'
ws_r['A1'].font = title_font
ws_r['A1'].fill = title_fill
ws_r['A1'].alignment = center
ws_r.row_dimensions[1].height = 24

raw_headers = ['연월', 'HS코드', '수출액 (USD, FOB)', '수출중량 (kg)', '단가 ($/kg)', '데이터출처']
for i, h in enumerate(raw_headers):
    c = ws_r.cell(row=2, column=i+1, value=h)
    c.font = header_font
    c.fill = header_fill
    c.alignment = center
    c.border = border

for i, r in enumerate(records):
    rr = 3 + i
    ws_r.cell(row=rr, column=1, value=r['period']).alignment = center
    ws_r.cell(row=rr, column=2, value='2841.90').alignment = center
    ws_r.cell(row=rr, column=3, value=r['value_usd']).number_format = '#,##0'
    ws_r.cell(row=rr, column=3).alignment = right
    ws_r.cell(row=rr, column=4, value=r['wgt_ton']*1000).number_format = '#,##0'
    ws_r.cell(row=rr, column=4).alignment = right
    ws_r.cell(row=rr, column=5, value=r['unit_price']/1000 if r['unit_price'] else None).number_format = '#,##0.00'
    ws_r.cell(row=rr, column=5).alignment = right
    ws_r.cell(row=rr, column=6, value='UN Comtrade').alignment = center
    for c in range(1, 7):
        ws_r.cell(row=rr, column=c).border = border
        ws_r.cell(row=rr, column=c).font = Font(name='맑은 고딕', size=10)

widths_r = [10, 10, 20, 16, 14, 16]
for i, w in enumerate(widths_r):
    ws_r.column_dimensions[get_column_letter(i+1)].width = w

# ============ Sheet 0: README ============
ws_h = wb.create_sheet('README', 0)
ws_h.sheet_view.showGridLines = False
ws_h.merge_cells('B2:H2')
ws_h['B2'] = '수출입통계 트래커 — Prototype v0.1'
ws_h['B2'].font = Font(name='맑은 고딕', size=18, bold=True, color=WHITE)
ws_h['B2'].fill = title_fill
ws_h['B2'].alignment = center
ws_h.row_dimensions[2].height = 36

readme_lines = [
    ('대상 종목', '에코프로비엠 (247540)'),
    ('대상 HS코드', '2841.90 — 산의 금속산염류 (양극재 분류)'),
    ('데이터 출처', 'UN Comtrade Public Preview API (한국 -> 전세계 월별 수출, FOB USD)'),
    ('데이터 기간', f"{records[0]['period']} ~ {records[-1]['period']}  ({len(records)}개월)"),
    ('생성일', datetime.now().strftime('%Y-%m-%d %H:%M')),
    ('', ''),
    ('[ 시트 구성 ]', ''),
    ('1. Dashboard', '에코프로비엠 종목 대시보드 — 정보박스 + KPI + 시계열 표 + 차트 2개'),
    ('2. Universe', '시총 1,500억 이상 TOP500 종목 (현재 매핑 완료: 1종목)'),
    ('3. HS_Mapping', '종목 <-> HS코드 매핑 마스터 (현재 1건)'),
    ('4. Raw_Data', 'UN Comtrade 원본 데이터'),
    ('', ''),
    ('[ 정식 버전에서 추가될 것 ]', ''),
    ('* 종목 드롭다운', 'Dashboard 시트에서 종목 선택 -> 차트 자동 변경'),
    ('* 데이터 신선도', '관세청 OpenAPI 연동 시 1~2개월 내 최신 데이터 (Comtrade는 6~12개월 지연)'),
    ('* 자동 갱신', '매월 16일 자동 실행, 새 데이터 fetch + 엑셀 갱신'),
    ('', ''),
    ('[ 한계 ]', 'UN Comtrade 공개 preview는 최신 데이터가 16개월 정도 지연됨. 관세청 API 연동 시 해결.'),
]
for i, (k, v) in enumerate(readme_lines):
    r = 4 + i
    ws_h.cell(row=r, column=2, value=k).font = Font(name='맑은 고딕', size=11, bold=True, color=NAVY)
    ws_h.cell(row=r, column=2).alignment = left
    ws_h.merge_cells(start_row=r, start_column=3, end_row=r, end_column=8)
    ws_h.cell(row=r, column=3, value=v).font = Font(name='맑은 고딕', size=11)
    ws_h.cell(row=r, column=3).alignment = left

ws_h.column_dimensions['A'].width = 2
ws_h.column_dimensions['B'].width = 24
for c in 'CDEFGH':
    ws_h.column_dimensions[c].width = 14

os.makedirs('output', exist_ok=True)
out = 'output/수출입통계_Prototype_에코프로비엠.xlsx'
wb.save(out)
print(f'생성 완료: {out}')
print(f'파일 크기: {os.path.getsize(out)/1024:.1f} KB')
print(f'시트: {wb.sheetnames}')
