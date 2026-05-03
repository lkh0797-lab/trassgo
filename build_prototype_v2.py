"""에코프로비엠 v2 prototype — 관세청 시군구 데이터 (청주+포항 HS 284190)"""
import sys, io, json, os
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import LineChart, BarChart, Reference
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import ColorScaleRule
from datetime import datetime

# 데이터 로드 - v2 (관세청)
with open('data/raw/customs_284190_v2.json', 'r', encoding='utf-8') as f:
    v2_raw = json.load(f)

# 청주시 + 포항시만 추출, 월별 집계
def aggregate(records, sgg_filter):
    out = {}
    for r in records:
        if sgg_filter not in r['sggNm']: continue
        p = r['priodTitle']
        out[p] = {
            'sgg': r['sggNm'],
            'exp_kusd': int(r['expUsdAmt'].replace(',', '')) if r['expUsdAmt'] else 0,
            'exp_cnt': int(r['expCnt'].replace(',', '')) if r['expCnt'] else 0,
            'imp_kusd': int(r['impUsdAmt'].replace(',', '')) if r['impUsdAmt'] else 0,
        }
    return out

cheongju = aggregate(v2_raw['chungbuk'], '청주시')
pohang = aggregate(v2_raw['gyeongbuk'], '포항시')

all_periods = sorted(set(cheongju.keys()) | set(pohang.keys()))

# 합계 계산
combined = []
for p in all_periods:
    cj = cheongju.get(p, {}).get('exp_kusd', 0)
    ph = pohang.get(p, {}).get('exp_kusd', 0)
    combined.append({
        'period': p,
        'cheongju_kusd': cj,
        'pohang_kusd': ph,
        'total_kusd': cj + ph,
        'cheongju_cnt': cheongju.get(p, {}).get('exp_cnt', 0),
        'pohang_cnt': pohang.get(p, {}).get('exp_cnt', 0),
    })

# YoY/MoM
for i, r in enumerate(combined):
    r['mom'] = r['total_kusd'] / combined[i-1]['total_kusd'] - 1 if i >= 1 and combined[i-1]['total_kusd'] else None
    r['yoy'] = r['total_kusd'] / combined[i-12]['total_kusd'] - 1 if i >= 12 and combined[i-12]['total_kusd'] else None

# v1 데이터 로드 (UN Comtrade)
with open('data/raw/comtrade_284190_korea_export.json', 'r', encoding='utf-8') as f:
    v1_raw = json.load(f)
v1_by_period = {r['period'][:4]+'.'+r['period'][4:]: r.get('primaryValue', 0)/1000 for r in v1_raw}  # USD → kUSD

# 스타일
NAVY = '1F3864'
LIGHT_BLUE = 'D9E1F2'
GRAY = 'F2F2F2'
WHITE = 'FFFFFF'
RED = 'C00000'
GREEN = '548235'

title_font = Font(name='맑은 고딕', size=14, bold=True, color=WHITE)
title_fill = PatternFill('solid', fgColor=NAVY)
header_font = Font(name='맑은 고딕', size=10, bold=True, color=WHITE)
header_fill = PatternFill('solid', fgColor=NAVY)
sub_fill = PatternFill('solid', fgColor=LIGHT_BLUE)
gray_fill = PatternFill('solid', fgColor=GRAY)
center = Alignment(horizontal='center', vertical='center')
right = Alignment(horizontal='right', vertical='center')
left = Alignment(horizontal='left', vertical='center')
thin = Side(style='thin', color='BFBFBF')
border = Border(left=thin, right=thin, top=thin, bottom=thin)

wb = openpyxl.Workbook()

# ============ Sheet 1: Dashboard ============
ws = wb.active
ws.title = 'Dashboard'
ws.sheet_view.showGridLines = False

ws.merge_cells('B2:K2')
ws['B2'] = '수출입통계 트래커 v2 — 에코프로비엠 (관세청 시군구별 데이터)'
ws['B2'].font = title_font
ws['B2'].fill = title_fill
ws['B2'].alignment = center
ws.row_dimensions[2].height = 28

info_labels = [
    ('종목명', '에코프로비엠'),
    ('종목코드', '247540'),
    ('시장', 'KOSDAQ'),
    ('시총(억원)', '93,847'),
    ('주력제품', '하이니켈 NCM/NCA 양극활물질'),
    ('HS코드', '2841.90 (산의 금속산염류)'),
    ('생산거점', '충청북도 청주시 오창공장 + 경상북도 포항시 영일만공장'),
    ('추정 공식', '청주시 양극재 수출 + 포항시 양극재 수출 (HS 284190)'),
    ('데이터 출처', '관세청 OpenAPI - 시군구별 품목별 수출입실적'),
    ('데이터 기간', f"{all_periods[0]} ~ {all_periods[-1]} ({len(all_periods)}개월)"),
    ('단위 주의', '응답값은 천 USD ($1,000) 단위'),
]
start_row = 4
for i, (k, v) in enumerate(info_labels):
    r = start_row + i
    ws.cell(row=r, column=2, value=k).font = Font(name='맑은 고딕', size=10, bold=True, color=NAVY)
    ws.cell(row=r, column=2).fill = sub_fill
    ws.cell(row=r, column=2).alignment = left
    ws.cell(row=r, column=2).border = border
    ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=6)
    ws.cell(row=r, column=3, value=v).font = Font(name='맑은 고딕', size=10)
    ws.cell(row=r, column=3).alignment = left
    for c in range(3, 7):
        ws.cell(row=r, column=c).border = border

# KPI
latest = combined[-1]
peak = max(combined, key=lambda x: x['total_kusd'])

kpis = [
    ('최근월', latest['period']),
    ('청주 수출', f"${latest['cheongju_kusd']/1000:,.1f}M"),
    ('포항 수출', f"${latest['pohang_kusd']/1000:,.1f}M"),
    ('합계 (에코프로비엠 추정)', f"${latest['total_kusd']/1000:,.1f}M"),
    ('YoY', f"{latest['yoy']*100:+.1f}%" if latest['yoy'] is not None else 'N/A'),
    ('MoM', f"{latest['mom']*100:+.1f}%" if latest['mom'] is not None else 'N/A'),
    ('역대 피크', f"${peak['total_kusd']/1000:,.1f}M  ({peak['period']})"),
    ('피크 대비', f"{(latest['total_kusd']/peak['total_kusd'] - 1)*100:+.1f}%"),
]
for i, (k, v) in enumerate(kpis):
    r = start_row + i
    ws.cell(row=r, column=8, value=k).font = Font(name='맑은 고딕', size=10, bold=True, color=NAVY)
    ws.cell(row=r, column=8).fill = sub_fill
    ws.cell(row=r, column=8).alignment = left
    ws.cell(row=r, column=8).border = border
    ws.merge_cells(start_row=r, start_column=9, end_row=r, end_column=11)
    ws.cell(row=r, column=9, value=v).font = Font(name='맑은 고딕', size=11, bold=True)
    ws.cell(row=r, column=9).alignment = right
    for c in range(9, 12):
        ws.cell(row=r, column=c).border = border

# 시계열 표 (합계만)
table_start = start_row + len(info_labels) + 2
ws.cell(row=table_start, column=2, value='월별 수출 시계열 (단위: 천 USD)').font = Font(name='맑은 고딕', size=12, bold=True, color=NAVY)

th_row = table_start + 1
headers = ['연월', '수출액 (천USD)', 'MoM', 'YoY']
for i, h in enumerate(headers):
    c = ws.cell(row=th_row, column=2+i, value=h)
    c.font = header_font
    c.fill = header_fill
    c.alignment = center
    c.border = border

for i, r in enumerate(combined):
    rr = th_row + 1 + i
    ws.cell(row=rr, column=2, value=r['period']).alignment = center
    ws.cell(row=rr, column=3, value=r['total_kusd']).number_format = '#,##0'
    ws.cell(row=rr, column=3).alignment = right
    ws.cell(row=rr, column=3).font = Font(name='맑은 고딕', size=10, bold=True)
    ws.cell(row=rr, column=4, value=r['mom']).number_format = '+0.0%;-0.0%;-'
    ws.cell(row=rr, column=4).alignment = right
    ws.cell(row=rr, column=5, value=r['yoy']).number_format = '+0.0%;-0.0%;-'
    ws.cell(row=rr, column=5).alignment = right
    if i % 2 == 0:
        for c in range(2, 6):
            if c != 3: ws.cell(row=rr, column=c).fill = gray_fill
    for c in range(2, 6):
        ws.cell(row=rr, column=c).border = border
        if c != 3:
            ws.cell(row=rr, column=c).font = Font(name='맑은 고딕', size=10)

last_data_row = th_row + len(combined)
ws.conditional_formatting.add(
    f'E{th_row+1}:E{last_data_row}',
    ColorScaleRule(start_type='num', start_value=-0.5, start_color='F8696B',
                   mid_type='num', mid_value=0, mid_color='FFEB84',
                   end_type='num', end_value=0.5, end_color='63BE7B'))

ws.column_dimensions['A'].width = 2
ws.column_dimensions['B'].width = 14
ws.column_dimensions['C'].width = 18
ws.column_dimensions['D'].width = 12
ws.column_dimensions['E'].width = 12
for c in ['F', 'G', 'H']:
    ws.column_dimensions[c].width = 14
ws.column_dimensions['I'].width = 12
ws.column_dimensions['J'].width = 12
ws.column_dimensions['K'].width = 12

# 차트 1: 수출액 추이
chart1 = LineChart()
chart1.title = '월별 수출액 추이 (천 USD)'
chart1.y_axis.title = '천 USD'
chart1.height = 9
chart1.width = 22
data_ref = Reference(ws, min_col=3, min_row=th_row, max_col=3, max_row=last_data_row)
cat_ref = Reference(ws, min_col=2, min_row=th_row+1, max_row=last_data_row)
chart1.add_data(data_ref, titles_from_data=True)
chart1.set_categories(cat_ref)
chart1.legend = None
ws.add_chart(chart1, 'G16')

# 차트 2: YoY
chart2 = BarChart()
chart2.type = 'col'
chart2.title = 'YoY 증감률'
chart2.height = 9
chart2.width = 22
yoy_ref = Reference(ws, min_col=5, min_row=th_row, max_col=5, max_row=last_data_row)
chart2.add_data(yoy_ref, titles_from_data=True)
chart2.set_categories(cat_ref)
chart2.legend = None
ws.add_chart(chart2, 'G35')

# ============ Sheet: Cheongju_Detail (청주시 상세) ============
def make_detail_sheet(name, title, records_dict):
    ws_d = wb.create_sheet(name)
    ws_d.sheet_view.showGridLines = False
    ws_d.merge_cells('A1:G1')
    ws_d['A1'] = title
    ws_d['A1'].font = title_font
    ws_d['A1'].fill = title_fill
    ws_d['A1'].alignment = center
    ws_d.row_dimensions[1].height = 24

    headers = ['연월', '시군구', '수출액 (천USD)', '수출건수', '수입액 (천USD)', 'MoM', 'YoY']
    for i, h in enumerate(headers):
        c = ws_d.cell(row=2, column=i+1, value=h)
        c.font = header_font
        c.fill = header_fill
        c.alignment = center
        c.border = border

    sorted_periods = sorted(records_dict.keys())
    vals = [records_dict[p]['exp_kusd'] for p in sorted_periods]
    for i, p in enumerate(sorted_periods):
        r = records_dict[p]
        rr = 3 + i
        mom = vals[i]/vals[i-1] - 1 if i >= 1 and vals[i-1] else None
        yoy = vals[i]/vals[i-12] - 1 if i >= 12 and vals[i-12] else None
        ws_d.cell(row=rr, column=1, value=p).alignment = center
        ws_d.cell(row=rr, column=2, value=r['sgg']).alignment = left
        ws_d.cell(row=rr, column=3, value=r['exp_kusd']).number_format = '#,##0'
        ws_d.cell(row=rr, column=3).alignment = right
        ws_d.cell(row=rr, column=4, value=r['exp_cnt']).number_format = '#,##0'
        ws_d.cell(row=rr, column=4).alignment = right
        ws_d.cell(row=rr, column=5, value=r['imp_kusd']).number_format = '#,##0'
        ws_d.cell(row=rr, column=5).alignment = right
        ws_d.cell(row=rr, column=6, value=mom).number_format = '+0.0%;-0.0%;-'
        ws_d.cell(row=rr, column=6).alignment = right
        ws_d.cell(row=rr, column=7, value=yoy).number_format = '+0.0%;-0.0%;-'
        ws_d.cell(row=rr, column=7).alignment = right
        if i % 2 == 0:
            for c in range(1, 8):
                ws_d.cell(row=rr, column=c).fill = gray_fill
        for c in range(1, 8):
            ws_d.cell(row=rr, column=c).border = border
            ws_d.cell(row=rr, column=c).font = Font(name='맑은 고딕', size=10)

    last = 2 + len(sorted_periods)
    ws_d.conditional_formatting.add(
        f'G3:G{last}',
        ColorScaleRule(start_type='num', start_value=-0.5, start_color='F8696B',
                       mid_type='num', mid_value=0, mid_color='FFEB84',
                       end_type='num', end_value=0.5, end_color='63BE7B'))

    chart = LineChart()
    chart.title = title
    chart.height = 9
    chart.width = 22
    data_ref = Reference(ws_d, min_col=3, min_row=2, max_col=3, max_row=last)
    cat_ref = Reference(ws_d, min_col=1, min_row=3, max_row=last)
    chart.add_data(data_ref, titles_from_data=True)
    chart.set_categories(cat_ref)
    chart.legend = None
    ws_d.add_chart(chart, 'I3')

    widths = [12, 18, 16, 12, 14, 10, 10]
    for i, w in enumerate(widths):
        ws_d.column_dimensions[get_column_letter(i+1)].width = w

make_detail_sheet('청주시_상세', '충청북도 청주시 양극재 수출 (HS 284190)', cheongju)
make_detail_sheet('포항시_상세', '경상북도 포항시 양극재 수출 (HS 284190)', pohang)

# ============ Sheet: HS_Mapping (업데이트) ============
ws_m = wb.create_sheet('HS_Mapping')
ws_m.sheet_view.showGridLines = False
ws_m.merge_cells('A1:I1')
ws_m['A1'] = '종목 ↔ HS코드 ↔ 시군구 매핑 마스터 (v2)'
ws_m['A1'].font = title_font
ws_m['A1'].fill = title_fill
ws_m['A1'].alignment = center
ws_m.row_dimensions[1].height = 24

map_headers = ['종목코드', '종목명', 'HS6', 'HS 한글품명', '주력제품', '시도코드', '시군구', '추정 점유율 메모', '검토상태']
for i, h in enumerate(map_headers):
    c = ws_m.cell(row=2, column=i+1, value=h)
    c.font = header_font
    c.fill = header_fill
    c.alignment = center
    c.border = border

mapping_rows = [
    ['247540', '에코프로비엠', '284190', '산의 금속산염류 (양극재)', 'NCM/NCA 양극활물질', '43', '청주시', '단독 (오창공장)', '확정'],
    ['247540', '에코프로비엠', '284190', '산의 금속산염류 (양극재)', 'NCM/NCA 양극활물질', '47', '포항시', '포스코퓨처엠과 공유 → 비중 추정 필요', '검토중'],
]
for i, mr in enumerate(mapping_rows):
    rr = 3 + i
    for j, v in enumerate(mr):
        c = ws_m.cell(row=rr, column=j+1, value=v)
        c.font = Font(name='맑은 고딕', size=10)
        c.border = border
        c.alignment = center if j in (0, 2, 5, 8) else left

widths_m = [10, 14, 8, 22, 20, 9, 12, 32, 11]
for i, w in enumerate(widths_m):
    ws_m.column_dimensions[get_column_letter(i+1)].width = w

# ============ Sheet: Raw_Data ============
ws_r = wb.create_sheet('Raw_Data')
ws_r.sheet_view.showGridLines = False
ws_r.merge_cells('A1:H1')
ws_r['A1'] = '관세청 시군구별 품목별 수출입실적 - 충북 + 경북 HS 284190 36개월'
ws_r['A1'].font = title_font
ws_r['A1'].fill = title_fill
ws_r['A1'].alignment = center
ws_r.row_dimensions[1].height = 24

raw_headers = ['연월', '시도', '시군구', '수출건수', '수출액(천USD)', '수입건수', '수입액(천USD)', '무역수지(천USD)']
for i, h in enumerate(raw_headers):
    c = ws_r.cell(row=2, column=i+1, value=h)
    c.font = header_font
    c.fill = header_fill
    c.alignment = center
    c.border = border

all_records = []
for r in v2_raw['chungbuk']:
    all_records.append(('충청북도', r))
for r in v2_raw['gyeongbuk']:
    all_records.append(('경상북도', r))

# 정렬: 연월 → 시도 → 시군구
all_records.sort(key=lambda x: (x[1]['priodTitle'], x[0], x[1]['sggNm']))

for i, (sido, r) in enumerate(all_records):
    rr = 3 + i
    ws_r.cell(row=rr, column=1, value=r['priodTitle']).alignment = center
    ws_r.cell(row=rr, column=2, value=sido).alignment = center
    ws_r.cell(row=rr, column=3, value=r['sggNm']).alignment = left
    ws_r.cell(row=rr, column=4, value=int(r['expCnt'].replace(',', '')) if r['expCnt'] else 0).number_format = '#,##0'
    ws_r.cell(row=rr, column=5, value=int(r['expUsdAmt'].replace(',', '')) if r['expUsdAmt'] else 0).number_format = '#,##0'
    ws_r.cell(row=rr, column=6, value=int(r['impCnt'].replace(',', '')) if r['impCnt'] else 0).number_format = '#,##0'
    ws_r.cell(row=rr, column=7, value=int(r['impUsdAmt'].replace(',', '')) if r['impUsdAmt'] else 0).number_format = '#,##0'
    ws_r.cell(row=rr, column=8, value=int(r['cmtrBlncAmt'].replace(',', '')) if r['cmtrBlncAmt'] else 0).number_format = '#,##0'
    for c in range(4, 9):
        ws_r.cell(row=rr, column=c).alignment = right
    for c in range(1, 9):
        ws_r.cell(row=rr, column=c).border = border
        ws_r.cell(row=rr, column=c).font = Font(name='맑은 고딕', size=9)

widths_r = [11, 10, 18, 10, 14, 10, 14, 14]
for i, w in enumerate(widths_r):
    ws_r.column_dimensions[get_column_letter(i+1)].width = w
ws_r.freeze_panes = 'A3'

# ============ Sheet 0: README ============
ws_h = wb.create_sheet('README', 0)
ws_h.sheet_view.showGridLines = False
ws_h.merge_cells('B2:H2')
ws_h['B2'] = '수출입통계 트래커 — Prototype v2 (관세청 실데이터)'
ws_h['B2'].font = Font(name='맑은 고딕', size=18, bold=True, color=WHITE)
ws_h['B2'].fill = title_fill
ws_h['B2'].alignment = center
ws_h.row_dimensions[2].height = 36

readme_lines = [
    ('대상 종목', '에코프로비엠 (247540)'),
    ('대상 HS코드', '284190 (산의 금속산염류 - 양극재)'),
    ('데이터 출처', '관세청 OpenAPI - 시군구별 품목별 수출입실적'),
    ('회사 추정 공식', '충북 청주시 + 경북 포항시 양극재 수출액 합산'),
    ('데이터 기간', f"{all_periods[0]} ~ {all_periods[-1]} ({len(all_periods)}개월)"),
    ('생성일', datetime.now().strftime('%Y-%m-%d %H:%M')),
    ('', ''),
    ('[ 시트 구성 ]', ''),
    ('1. Dashboard', '에코프로비엠 종목 대시보드 - KPI + 시계열 합계 + 차트'),
    ('2. 청주시_상세', '충북 청주시 양극재 수출 36개월 시계열 + 차트'),
    ('3. 포항시_상세', '경북 포항시 양극재 수출 36개월 시계열 + 차트 (포스코퓨처엠 합산)'),
    ('4. HS_Mapping', '종목-HS-시군구 매핑 마스터'),
    ('5. Raw_Data', '관세청 원본 데이터 (충북+경북 모든 시군구)'),
    ('', ''),
    ('[ 핵심 인사이트 ]', ''),
    ('* 청주공장 가동중단 신호', '2024-08 $4M까지 폭락 후 점진 회복 (피크 대비 -97%)'),
    ('* 포항공장 안정 운영', '$100~150M 박스권 유지 - 회사를 살리는 중'),
    ('* 합계 피크', '{} ${:,.0f}M -> 최근 ${:,.0f}M ({:+.0f}%)'.format(peak['period'], peak['total_kusd']/1000, latest['total_kusd']/1000, (latest['total_kusd']/peak['total_kusd']-1)*100)),
    ('* 데이터 신선도', f"최근월 {latest['period']} - 5월 1일 발표라면 2026-04분도 곧 가능"),
    ('', ''),
    ('[ 한계 및 정밀화 필요 ]', ''),
    ('* 포항 합산 노이즈', '포스코퓨처엠 양극재도 포항에 있음 -> 분리 필요 (사업보고서 매출비중 활용)'),
    ('* 청주 외 공장', '에코프로 충북 음성/진천 공장도 일부 양극재 - 추가 검토'),
    ('* 단위 명확화', '응답값은 천 USD ($1,000) - 표 헤더에 단위 표시'),
]
for i, (k, v) in enumerate(readme_lines):
    r = 4 + i
    ws_h.cell(row=r, column=2, value=k).font = Font(name='맑은 고딕', size=11, bold=True, color=NAVY)
    ws_h.cell(row=r, column=2).alignment = left
    ws_h.merge_cells(start_row=r, start_column=3, end_row=r, end_column=8)
    ws_h.cell(row=r, column=3, value=v).font = Font(name='맑은 고딕', size=11)
    ws_h.cell(row=r, column=3).alignment = left

ws_h.column_dimensions['A'].width = 2
ws_h.column_dimensions['B'].width = 26
for c in 'CDEFGH':
    ws_h.column_dimensions[c].width = 14

os.makedirs('output', exist_ok=True)
out = 'output/수출입통계_Prototype_v2_에코프로비엠_revised.xlsx'
wb.save(out)
print(f'생성 완료: {out}')
print(f'파일 크기: {os.path.getsize(out)/1024:.1f} KB')
print(f'시트: {wb.sheetnames}')
