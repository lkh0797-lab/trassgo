"""종목별 시트 자동 생성 엑셀 대시보드 빌더"""
from __future__ import annotations

from datetime import datetime
from pathlib import Path

import openpyxl
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from .customs_api import load_settings
from .mapper import (
    MappingRow, StockTimeseries, group_by_stock, load_mappings, load_sido_codes,
)


PROJECT_ROOT = Path(__file__).resolve().parent.parent
OUTPUT_DIR = PROJECT_ROOT / 'output'
UNIVERSE_FILE_NAME = load_settings()['data']['universe_file']

# 스타일
NAVY = '1F3864'
LIGHT_BLUE = 'D9E1F2'
GRAY = 'F2F2F2'
WHITE = 'FFFFFF'

TITLE_FONT = Font(name='맑은 고딕', size=14, bold=True, color=WHITE)
TITLE_FILL = PatternFill('solid', fgColor=NAVY)
HEADER_FONT = Font(name='맑은 고딕', size=10, bold=True, color=WHITE)
HEADER_FILL = PatternFill('solid', fgColor=NAVY)
SUB_FILL = PatternFill('solid', fgColor=LIGHT_BLUE)
GRAY_FILL = PatternFill('solid', fgColor=GRAY)
CENTER = Alignment(horizontal='center', vertical='center')
RIGHT = Alignment(horizontal='right', vertical='center')
LEFT = Alignment(horizontal='left', vertical='center')
THIN = Side(style='thin', color='BFBFBF')
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def _set_title(ws, cell_range: str, text: str, height: int = 26):
    ws.merge_cells(cell_range)
    first_cell = cell_range.split(':')[0]
    ws[first_cell] = text
    ws[first_cell].font = TITLE_FONT
    ws[first_cell].fill = TITLE_FILL
    ws[first_cell].alignment = CENTER
    row = int(''.join(c for c in first_cell if c.isdigit()))
    ws.row_dimensions[row].height = height


def _header_row(ws, row: int, headers: list[str], start_col: int = 1):
    for i, h in enumerate(headers):
        c = ws.cell(row=row, column=start_col + i, value=h)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.alignment = CENTER
        c.border = BORDER


def build_summary_sheet(wb, ts_by_stock: dict[str, StockTimeseries]):
    ws = wb.create_sheet('Summary')
    ws.sheet_view.showGridLines = False
    _set_title(ws, 'A1:K1', '📊 매핑 종목 요약 (Summary)')

    headers = ['종목코드', '종목명', '시도/시군구', 'HS코드', '최근월',
               '최근 수출액 (천USD)', 'YoY', 'MoM', '3M평균', '12M평균', '피크 대비']
    _header_row(ws, 2, headers)

    rr = 3
    for code, ts in sorted(ts_by_stock.items()):
        if not ts.by_period:
            continue
        sorted_p = sorted(ts.by_period.keys())
        latest = sorted_p[-1]
        slot = ts.by_period[latest]

        sgg_str = ', '.join(f"{m.sgg_keyword}({m.sido_code})" for m in ts.mappings)
        hs_str = ', '.join(sorted(set(m.hs_code for m in ts.mappings)))

        last_3 = [ts.by_period[p]['exp_kusd'] for p in sorted_p[-3:]]
        last_12 = [ts.by_period[p]['exp_kusd'] for p in sorted_p[-12:]]
        peak = max(ts.by_period[p]['exp_kusd'] for p in sorted_p) or 1
        vs_peak = slot['exp_kusd'] / peak - 1

        cells = [
            (code, CENTER), (ts.stock_name, LEFT), (sgg_str, LEFT), (hs_str, CENTER),
            (latest, CENTER), (slot['exp_kusd'], RIGHT), (slot['yoy'], RIGHT),
            (slot['mom'], RIGHT),
            (sum(last_3) / len(last_3), RIGHT), (sum(last_12) / len(last_12), RIGHT),
            (vs_peak, RIGHT),
        ]
        for i, (v, align) in enumerate(cells):
            c = ws.cell(row=rr, column=i + 1, value=v)
            c.font = Font(name='맑은 고딕', size=10)
            c.alignment = align
            c.border = BORDER
            if i in (5, 8, 9):
                c.number_format = '#,##0'
            elif i in (6, 7, 10):
                c.number_format = '+0.0%;-0.0%;-'
        rr += 1

    widths = [10, 16, 22, 12, 10, 18, 10, 10, 14, 14, 12]
    for i, w in enumerate(widths):
        ws.column_dimensions[get_column_letter(i + 1)].width = w


def build_stock_detail_sheet(wb, ts: StockTimeseries):
    sheet_name = f'{ts.stock_name}_상세'[:31]
    ws = wb.create_sheet(sheet_name)
    ws.sheet_view.showGridLines = False
    _set_title(ws, 'B2:K2', f'📈 {ts.stock_name} ({ts.stock_code}) — 수출 시계열')

    sgg_str = ', '.join(f"{m.sgg_keyword}({m.sido_code})" for m in ts.mappings)
    hs_str = ', '.join(f"{m.hs_code} ({m.hs_desc})" for m in ts.mappings)

    info_rows = [
        ('종목코드', ts.stock_code),
        ('종목명', ts.stock_name),
        ('HS코드', hs_str),
        ('생산거점 (시도/시군구)', sgg_str),
        ('데이터 단위', '천 USD ($1,000)'),
    ]
    for i, (k, v) in enumerate(info_rows):
        r = 4 + i
        ws.cell(row=r, column=2, value=k).font = Font(name='맑은 고딕', size=10, bold=True, color=NAVY)
        ws.cell(row=r, column=2).fill = SUB_FILL
        ws.cell(row=r, column=2).alignment = LEFT
        ws.cell(row=r, column=2).border = BORDER
        ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=6)
        ws.cell(row=r, column=3, value=v).font = Font(name='맑은 고딕', size=10)
        ws.cell(row=r, column=3).alignment = LEFT
        for c in range(3, 7):
            ws.cell(row=r, column=c).border = BORDER

    sorted_p = sorted(ts.by_period.keys())
    latest_slot = ts.by_period[sorted_p[-1]] if sorted_p else None

    if latest_slot:
        peak_p = max(sorted_p, key=lambda p: ts.by_period[p]['exp_kusd'])
        peak_v = ts.by_period[peak_p]['exp_kusd']
        last_3 = sum(ts.by_period[p]['exp_kusd'] for p in sorted_p[-3:]) / 3
        last_12 = sum(ts.by_period[p]['exp_kusd'] for p in sorted_p[-12:]) / min(12, len(sorted_p))

        kpis = [
            ('최근월', sorted_p[-1]),
            ('최근 수출액', f"${latest_slot['exp_kusd']/1000:,.1f}M"),
            ('YoY', f"{latest_slot['yoy']*100:+.1f}%" if latest_slot['yoy'] is not None else 'N/A'),
            ('MoM', f"{latest_slot['mom']*100:+.1f}%" if latest_slot['mom'] is not None else 'N/A'),
            ('3개월 평균', f"${last_3/1000:,.1f}M"),
            ('12개월 평균', f"${last_12/1000:,.1f}M"),
            ('역대 피크', f"${peak_v/1000:,.1f}M ({peak_p})"),
            ('피크 대비', f"{(latest_slot['exp_kusd']/peak_v - 1)*100:+.1f}%" if peak_v else 'N/A'),
        ]
        for i, (k, v) in enumerate(kpis):
            r = 4 + i
            ws.cell(row=r, column=8, value=k).font = Font(name='맑은 고딕', size=10, bold=True, color=NAVY)
            ws.cell(row=r, column=8).fill = SUB_FILL
            ws.cell(row=r, column=8).alignment = LEFT
            ws.cell(row=r, column=8).border = BORDER
            ws.merge_cells(start_row=r, start_column=9, end_row=r, end_column=11)
            ws.cell(row=r, column=9, value=v).font = Font(name='맑은 고딕', size=11, bold=True)
            ws.cell(row=r, column=9).alignment = RIGHT
            for c in range(9, 12):
                ws.cell(row=r, column=c).border = BORDER

    table_start = 4 + max(len(info_rows), 8) + 2
    ws.cell(row=table_start, column=2, value='월별 수출 시계열 (단위: 천 USD)').font = Font(
        name='맑은 고딕', size=12, bold=True, color=NAVY)

    th_row = table_start + 1
    headers = ['연월', '수출액 (천USD)', '수출건수', 'MoM', 'YoY']
    _header_row(ws, th_row, headers, start_col=2)

    for i, p in enumerate(sorted_p):
        slot = ts.by_period[p]
        rr = th_row + 1 + i
        ws.cell(row=rr, column=2, value=p).alignment = CENTER
        ws.cell(row=rr, column=3, value=slot['exp_kusd']).number_format = '#,##0'
        ws.cell(row=rr, column=3).alignment = RIGHT
        ws.cell(row=rr, column=3).font = Font(name='맑은 고딕', size=10, bold=True)
        ws.cell(row=rr, column=4, value=slot['exp_count']).number_format = '#,##0'
        ws.cell(row=rr, column=4).alignment = RIGHT
        ws.cell(row=rr, column=5, value=slot['mom']).number_format = '+0.0%;-0.0%;-'
        ws.cell(row=rr, column=5).alignment = RIGHT
        ws.cell(row=rr, column=6, value=slot['yoy']).number_format = '+0.0%;-0.0%;-'
        ws.cell(row=rr, column=6).alignment = RIGHT
        if i % 2 == 0:
            for c in range(2, 7):
                if c != 3:
                    ws.cell(row=rr, column=c).fill = GRAY_FILL
        for c in range(2, 7):
            ws.cell(row=rr, column=c).border = BORDER
            if c != 3:
                ws.cell(row=rr, column=c).font = Font(name='맑은 고딕', size=10)

    last_data_row = th_row + len(sorted_p)
    if sorted_p:
        ws.conditional_formatting.add(
            f'F{th_row+1}:F{last_data_row}',
            ColorScaleRule(start_type='num', start_value=-0.5, start_color='F8696B',
                           mid_type='num', mid_value=0, mid_color='FFEB84',
                           end_type='num', end_value=0.5, end_color='63BE7B'))

        chart1 = LineChart()
        chart1.title = '월별 수출액 추이 (천 USD)'
        chart1.height = 9
        chart1.width = 22
        chart1.legend = None
        data_ref = Reference(ws, min_col=3, min_row=th_row, max_col=3, max_row=last_data_row)
        cat_ref = Reference(ws, min_col=2, min_row=th_row + 1, max_row=last_data_row)
        chart1.add_data(data_ref, titles_from_data=True)
        chart1.set_categories(cat_ref)
        ws.add_chart(chart1, 'H16')

        chart2 = BarChart()
        chart2.type = 'col'
        chart2.title = 'YoY 증감률'
        chart2.height = 9
        chart2.width = 22
        chart2.legend = None
        yoy_ref = Reference(ws, min_col=6, min_row=th_row, max_col=6, max_row=last_data_row)
        chart2.add_data(yoy_ref, titles_from_data=True)
        chart2.set_categories(cat_ref)
        ws.add_chart(chart2, 'H35')

    widths = ['A:2', 'B:14', 'C:18', 'D:12', 'E:12', 'F:12', 'G:14', 'H:16', 'I:14', 'J:14', 'K:14']
    for w in widths:
        col, val = w.split(':')
        ws.column_dimensions[col].width = int(val)


def build_universe_sheet(wb, mappings: list[MappingRow]):
    ws = wb.create_sheet('Universe')
    ws.sheet_view.showGridLines = False

    universe_path = PROJECT_ROOT / UNIVERSE_FILE_NAME
    if not universe_path.exists():
        ws['A1'] = f'Universe 파일 없음: {UNIVERSE_FILE_NAME}'
        return

    src_wb = openpyxl.load_workbook(universe_path, data_only=True)
    src_ws = src_wb.active

    _set_title(ws, 'A1:L1', f'📋 Universe — {UNIVERSE_FILE_NAME}')

    new_headers = ['순위', '종목코드', '종목명', '시장', '시총(억원)', '현재가', '거래량',
                   '상장주식수(천주)', 'PER', 'ROE', 'HS 매핑상태', '편입여부']
    _header_row(ws, 2, new_headers)

    mapped_codes = {m.stock_code: m for m in mappings}

    for src_row in src_ws.iter_rows(min_row=3, values_only=True):
        if not src_row or src_row[0] is None:
            continue
        target_row = ws.max_row + 1
        for i, v in enumerate(src_row[:10]):
            c = ws.cell(row=target_row, column=i + 1, value=v)
            c.font = Font(name='맑은 고딕', size=9)
            c.border = BORDER
            c.alignment = CENTER if i in (0, 1, 3) else (RIGHT if i in (4, 5, 6, 7, 8, 9) else LEFT)

        code = str(src_row[1]) if src_row[1] else ''
        if code in mapped_codes:
            m = mapped_codes[code]
            for c_idx in range(1, 11):
                ws.cell(row=target_row, column=c_idx).fill = PatternFill('solid', fgColor='FFF2CC')
            mc = ws.cell(row=target_row, column=11, value=f'완료 (HS {m.hs_code})')
            mc.font = Font(name='맑은 고딕', size=9, bold=True, color='006100')
            mc.fill = PatternFill('solid', fgColor='C6EFCE')
            mc.border = BORDER
            mc.alignment = CENTER
            ec = ws.cell(row=target_row, column=12, value='편입')
            ec.font = Font(name='맑은 고딕', size=9, bold=True, color='006100')
            ec.fill = PatternFill('solid', fgColor='C6EFCE')
            ec.border = BORDER
            ec.alignment = CENTER
        else:
            mc = ws.cell(row=target_row, column=11, value='미매핑')
            mc.font = Font(name='맑은 고딕', size=9, color='808080')
            mc.border = BORDER
            mc.alignment = CENTER
            ec = ws.cell(row=target_row, column=12, value='-')
            ec.alignment = CENTER
            ec.border = BORDER

    widths = [6, 10, 18, 8, 12, 11, 13, 14, 8, 8, 22, 10]
    for i, w in enumerate(widths):
        ws.column_dimensions[get_column_letter(i + 1)].width = w
    ws.freeze_panes = 'A3'


def build_mapping_sheet(wb, mappings: list[MappingRow]):
    ws = wb.create_sheet('HS_Mapping')
    ws.sheet_view.showGridLines = False
    _set_title(ws, 'A1:J1', '🔗 종목 ↔ HS ↔ 시군구 매핑 마스터')
    headers = ['종목코드', '종목명', 'HS6', 'HS 한글품명', '시도', '시군구', '가중치', '신뢰도', '메모', '추가일']
    _header_row(ws, 2, headers)

    sido_map = load_sido_codes()
    for i, m in enumerate(mappings):
        rr = 3 + i
        sido = sido_map.get(m.sido_code, m.sido_code)
        cells = [m.stock_code, m.stock_name, m.hs_code, m.hs_desc,
                 f'{sido}({m.sido_code})', m.sgg_keyword, m.weight, m.confidence,
                 m.note, m.added_date]
        for j, v in enumerate(cells):
            c = ws.cell(row=rr, column=j + 1, value=v)
            c.font = Font(name='맑은 고딕', size=10)
            c.border = BORDER
            c.alignment = CENTER if j in (0, 2, 5, 6, 7, 9) else LEFT
        if i % 2 == 0:
            for c_idx in range(1, 11):
                ws.cell(row=rr, column=c_idx).fill = GRAY_FILL

    widths = [10, 16, 8, 24, 14, 12, 8, 10, 32, 12]
    for i, w in enumerate(widths):
        ws.column_dimensions[get_column_letter(i + 1)].width = w


def build_readme_sheet(wb, ts_by_stock: dict[str, StockTimeseries], yymm_start: str, yymm_end: str):
    ws = wb.create_sheet('README', 0)
    ws.sheet_view.showGridLines = False
    _set_title(ws, 'B2:H2', '📘 수출입통계 트래커', height=36)

    n_stocks = len([t for t in ts_by_stock.values() if t.by_period])
    lines = [
        ('생성일', datetime.now().strftime('%Y-%m-%d %H:%M')),
        ('데이터 기간', f'{yymm_start} ~ {yymm_end}'),
        ('매핑 종목 수', str(n_stocks)),
        ('데이터 출처', '관세청 OpenAPI - 시군구별 품목별 수출입실적'),
        ('데이터 단위', '천 USD ($1,000)'),
        ('', ''),
        ('[ 시트 구성 ]', ''),
        ('1. Summary', '매핑된 모든 종목의 최근월 요약'),
        ('2. {종목명}_상세', '종목별 상세 — 시계열, KPI, 차트'),
        ('3. Universe', '시총 1500억+ 500종목 + 매핑 상태'),
        ('4. HS_Mapping', '종목 ↔ HS코드 ↔ 시군구 매핑 마스터'),
        ('', ''),
        ('[ 종목 추가 방법 ]', ''),
        ('대화형', '클로드에게 "{종목명} 추가해줘" — HS코드/시군구 추정 + 검증 지원'),
        ('직접 편집', 'data/master/stock_mapping.csv 한 줄 추가 후 runner 재실행'),
    ]
    for i, (k, v) in enumerate(lines):
        r = 4 + i
        ws.cell(row=r, column=2, value=k).font = Font(name='맑은 고딕', size=11, bold=True, color=NAVY)
        ws.cell(row=r, column=2).alignment = LEFT
        ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=8)
        ws.cell(row=r, column=3, value=v).font = Font(name='맑은 고딕', size=11)
        ws.cell(row=r, column=3).alignment = LEFT

    ws.column_dimensions['A'].width = 2
    ws.column_dimensions['B'].width = 22
    for c in 'CDEFGH':
        ws.column_dimensions[c].width = 14


def build_dashboard(ts_by_stock: dict[str, StockTimeseries], mappings: list[MappingRow],
                    yymm_start: str, yymm_end: str) -> Path:
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    build_summary_sheet(wb, ts_by_stock)
    for code, ts in sorted(ts_by_stock.items()):
        if ts.by_period:
            build_stock_detail_sheet(wb, ts)
    build_universe_sheet(wb, mappings)
    build_mapping_sheet(wb, mappings)
    build_readme_sheet(wb, ts_by_stock, yymm_start, yymm_end)

    OUTPUT_DIR.mkdir(exist_ok=True)
    yyyymm = datetime.now().strftime('%Y-%m')
    out_path = OUTPUT_DIR / f'수출입통계_트래커_{yyyymm}.xlsx'
    try:
        wb.save(out_path)
    except PermissionError:
        ts = datetime.now().strftime('%Y-%m-%d_%H%M%S')
        out_path = OUTPUT_DIR / f'수출입통계_트래커_{yyyymm}__{ts}.xlsx'
        wb.save(out_path)
        print(f'[WARN] 기본 파일이 열려있어 새 파일에 저장: {out_path.name}')
    return out_path


if __name__ == '__main__':
    from .mapper import build_stock_timeseries
    maps = load_mappings()
    ts = build_stock_timeseries(maps, '202304', '202603')
    out = build_dashboard(ts, maps, '202304', '202603')
    import sys, io
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
    print(f'Saved: {out}')
